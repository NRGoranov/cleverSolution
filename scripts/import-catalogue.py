"""Import kitchen equipment catalogue Excel into categorized products with images."""

from __future__ import annotations

import json
import re
import unicodedata
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_SOURCE = ROOT / "docs" / "Catalogue-Webpage_data-Equipment-03.08.2026.xlsx"
FALLBACK_SOURCE = Path(
    r"c:\Users\user\Downloads\Catalogue-Webpage_data-Equipment-03.08.2026.xlsx"
)
OUT_TS = ROOT / "data" / "products" / "kitchenware.ts"
OUT_JSON = ROOT / "data" / "products" / "catalogue-import.json"
IMAGES_ROOT = ROOT / "public" / "images" / "products"
EUR_TO_BGN = 1.95583

NS = {
    "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
}

SHEET_CONFIG = [
    {"subcategory": "combi-ovens", "label": "Конвектомати"},
    {"subcategory": "speed-ovens", "label": "Speed ovens"},
    {"subcategory": "ovens", "label": "Автоматични фурни"},
    {"subcategory": "sous-vide", "label": "Sous-Vide"},
    {"subcategory": "equipment", "label": "Оборудване"},
]

CYRILLIC_MAP = {
    "а": "a",
    "б": "b",
    "в": "v",
    "г": "g",
    "д": "d",
    "е": "e",
    "ж": "zh",
    "з": "z",
    "и": "i",
    "й": "y",
    "к": "k",
    "л": "l",
    "м": "m",
    "н": "n",
    "о": "o",
    "п": "p",
    "р": "r",
    "с": "s",
    "т": "t",
    "у": "u",
    "ф": "f",
    "х": "h",
    "ц": "ts",
    "ч": "ch",
    "ш": "sh",
    "щ": "sht",
    "ъ": "a",
    "ь": "",
    "ю": "yu",
    "я": "ya",
}


def slugify(text: str) -> str:
    text = text.strip().lower()
    out: list[str] = []
    for ch in text:
        if ch in CYRILLIC_MAP:
            out.append(CYRILLIC_MAP[ch])
            continue
        norm = unicodedata.normalize("NFKD", ch)
        norm = norm.encode("ascii", "ignore").decode("ascii")
        if norm:
            out.append(norm.lower())
    slug = re.sub(r"[^a-z0-9]+", "-", "".join(out)).strip("-")
    slug = re.sub(r"-{2,}", "-", slug)
    return slug[:80] or "product"


def clean_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r\n", "\n").replace("\r", "\n")
    text = text.replace("\uf0d8", "•").replace("×", "x")
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def preserves_leading_capitalization(text: str) -> bool:
    stripped = text.lstrip()
    if not stripped:
        return False
    if re.match(r"^i[A-Z]", stripped):
        return True
    if re.match(r"^[A-Z0-9]", stripped) and not stripped[0].islower():
        return True
    return False


def capitalize_sentences(text: str, *, preserve_leading: bool = False) -> str:
    """Capitalize the first letter of each sentence in prose text."""
    if not text:
        return text

    chars = list(text)
    at_sentence_start = True
    skip_leading = preserve_leading

    for i, ch in enumerate(chars):
        if at_sentence_start:
            if skip_leading:
                skip_leading = False
                if ch.isalpha() or ch.isdigit():
                    at_sentence_start = False
                continue
            if ch.islower():
                chars[i] = ch.upper()
                at_sentence_start = False
            elif ch == "i" and i + 1 < len(chars) and chars[i + 1].isupper():
                at_sentence_start = False
            elif ch.isalpha() or ch.isdigit():
                at_sentence_start = False
        if ch in ".!?…":
            at_sentence_start = True
        elif ch == "\n":
            at_sentence_start = True

    return "".join(chars)


INCOMPLETE_END_CHARS = set(",;:/\\|•–-")


def normalize_ending(text: str) -> str:
    """Drop incomplete trailing punctuation; keep . ! ? or no ending mark."""
    if not text:
        return text

    lines: list[str] = []
    for line in text.split("\n"):
        stripped = line.rstrip()
        while stripped and stripped[-1] in INCOMPLETE_END_CHARS:
            stripped = stripped[:-1].rstrip()
        # Collapse accidental spaces before a kept terminal mark
        stripped = re.sub(r"\s+([.!?…])$", r"\1", stripped)
        lines.append(stripped)
    return "\n".join(lines).strip()


def normalize_prose(text: str, *, preserve_leading: bool = False) -> str:
    text = capitalize_sentences(text, preserve_leading=preserve_leading)
    return normalize_ending(text)


def first_line(text: str) -> str:
    line = text.split("\n", 1)[0].strip()
    return line or text[:120]


def product_name(raw: str, model: str | None = None) -> str:
    line = first_line(raw)
    match = SPEC_LINE_RE.match(line.strip())
    if match:
        label = normalize_spec_label(match.group(1))
        value = match.group(2).strip()
        if is_spec_candidate(label, value) and label.lower() == "модел":
            # "iVario PRO® Model: 2-S +" → "iVario PRO® 2-S +"
            prefix = re.sub(
                r"\s*(?:Model|Модел)\s*$",
                "",
                match.group(1).strip(),
                flags=re.IGNORECASE,
            ).strip()
            if prefix:
                return f"{prefix} {value}".strip()
            return f"Модел {value}"
    return line


def extract_model(text: str) -> str | None:
    match = re.search(r"Модел:\s*([^\n(]+)", text, flags=re.IGNORECASE)
    if match:
        return match.group(1).strip()
    return None


def parse_price(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return round(float(value) * EUR_TO_BGN, 2) if float(value) > 0 else None
    text = str(value).strip().replace(",", ".")
    if not text or text in {"#DIV/0!", "#REF!", "-"}:
        return None
    try:
        amount = float(re.sub(r"[^\d.]+", "", text))
    except ValueError:
        return None
    return round(amount * EUR_TO_BGN, 2) if amount > 0 else None


def find_header_row(rows: list[tuple]) -> int | None:
    for idx, row in enumerate(rows):
        first = clean_text(row[0]) if row else ""
        second = clean_text(row[1]) if len(row) > 1 else ""
        if first in {"№", "No", "N"} and second.upper().startswith("ИЗДЕЛ"):
            return idx
    return None


def is_product_row(row: tuple) -> bool:
    if not row:
        return False
    number = row[0]
    if isinstance(number, (int, float)) and not isinstance(number, bool):
        return int(number) > 0
    if isinstance(number, str) and number.strip().isdigit():
        return True
    return False


# Spec lines in descriptions: short label + ":" + value with numbers
# e.g. "Капацитет: 6х 1/1-GN", "Тегло: 73.0 кг", "Модел: iCP XS"
SPEC_LINE_RE = re.compile(
    r"^[-\u2022\u25ba\*\s]*([^:\n]{1,60}):\s*(.+)$",
    re.MULTILINE,
)
SKIP_SPEC_LABELS = {
    "опционално",
    "основни характеристики",
}
SECTION_HEADER_MARKERS = (
    "разполага със",
    "включва",
    "основни характеристики",
)


def normalize_spec_label(label: str) -> str:
    text = label.strip(" -•►\t*")
    text = re.sub(r"\s+", " ", text)
    # Fold English / branded model headings into Модел
    if text.lower() == "model" or text.lower().endswith(" model"):
        return "Модел"
    return text


def is_spec_label(label: str) -> bool:
    if not label or label.lower() in SKIP_SPEC_LABELS:
        return False
    if re.match(r"^\d", label):
        return False
    words = re.findall(r"[\w\-®°\.]+", label, flags=re.UNICODE)
    # Short attribute labels (BG: "Вместимост", "Минимална височина на ваната")
    if not (1 <= len(words) <= 5):
        return False
    if len(label) > 55:
        return False
    # Skip branded headings / accessory list intros
    if "®" in label and label.lower() != "модел":
        return False
    lower = label.casefold()
    if "окомплектовка" in lower or "аксесоар" in lower:
        return False
    return True


DEGREE_IN_TEXT_RE = re.compile(r"(?<!N)[°º]|°C|°F", re.IGNORECASE)
TEMP_RANGE_RE = re.compile(
    r"от\s+\d+[°º][\w\s,.+-]*?до\s+\d+[°º][CFС]?(?:\s+max)?",
    re.IGNORECASE,
)


def line_has_degree(text: str) -> bool:
    cleaned = re.sub(r"N°\s*[\d.]+", "", text, flags=re.IGNORECASE)
    return bool(DEGREE_IN_TEXT_RE.search(cleaned))


def parse_degree_line(line: str) -> tuple[str, str] | None:
    text = line.strip().lstrip("-•►* ").strip().rstrip(".")
    if not text or not line_has_degree(text):
        return None

    # Accessory catalogue refs like "N° 60.71.643"
    if re.match(r"^\d+\s*x\s", text, re.IGNORECASE):
        return None

    # Feature acronym — still capture as a spec when it mentions degrees
    if "ΔT°" in text and not re.search(r"\d+\s*[°º]", text):
        return "ΔT° готвене", "Да"

    # Unit selector
    if "мерни единици" in text.casefold():
        units = re.search(r"([°ºCF]+(?:\s*[и\/]\s*[°ºCF]+)*)", text, re.I)
        return "Мерни единици", units.group(1) if units else "°C / °F"

    # Embedded feature bullets — keep descriptive clause around the degree value
    bullet_feature = re.match(r"^([A-Za-z0-9][^–-]{0,40})\s*[-–]\s*(.+)$", text)
    if bullet_feature and "patented" in text.casefold():
        label = bullet_feature.group(1).strip()
        value = bullet_feature.group(2).strip()
        return label, value

    range_match = TEMP_RANGE_RE.search(text)
    if range_match:
        value = range_match.group(0).strip()
        label = text[: range_match.start()].strip().rstrip(":,")
        if not label:
            label = "Температурен диапазон"
        elif len(label) > 55:
            words = label.split()
            label = " ".join(words[-5:]) if len(words) > 5 else label
        return label, value

    leading_range = re.match(
        r"^(?P<value>От\s+\d+[°º][\w\s,.+-]*?до\s+\d+[°º][CFС]?(?:\s+max)?)"
        r"(?:\s+(?P<rest>.+))?$",
        text,
        re.IGNORECASE,
    )
    if leading_range:
        value = leading_range.group("value").strip()
        rest = (leading_range.group("rest") or "").strip()
        if rest:
            value = f"{value} ({rest})"
        return "Температура", value

    stability = re.search(r"(±[\d,.]+\s*[°º][CFС]?)", text)
    if stability:
        return "Температурна точност", stability.group(1)

    if ":" in text:
        label, value = text.split(":", 1)
        label, value = label.strip(), value.strip()
        if label and value and line_has_degree(value):
            return label, value.rstrip(".")

    if re.search(r"\d+[°º]", text):
        return "Температура", text

    return None


BULLET_LINE_RE = re.compile(r"^[-\u2022\u25ba\*]\s+")
FEATURE_KEYWORDS = (
    "LED",
    "Steptronic",
    "ClimaPlus",
    "AUTOCLEAN",
    "Finishing",
    "MultiSteam",
    "CareControl",
    "iDensity",
    "iCooking",
    "iProduction",
    "iCare",
    "SPS",
)


def is_feature_bullet_line(line: str) -> bool:
    return bool(BULLET_LINE_RE.match(line.strip()))


def is_feature_section_header(text: str) -> bool:
    cleaned = text.strip().rstrip(".")
    if cleaned.endswith(":"):
        return True
    lower = cleaned.casefold()
    return any(lower.startswith(marker) for marker in SECTION_HEADER_MARKERS)


def should_drop_feature_line(text: str) -> bool:
    lower = text.casefold()
    return lower.startswith("опционално") or is_feature_section_header(text)


def parse_feature_line(line: str) -> tuple[str, str] | None:
    """Extract • / - bullet feature lines into label/value specs."""
    stripped = line.strip()
    if not is_feature_bullet_line(stripped):
        return None

    text = BULLET_LINE_RE.sub("", stripped).strip().rstrip(".;")
    if len(text) < 8 or should_drop_feature_line(text):
        return None

    if re.match(r"^\d+\s*x\s", text, re.IGNORECASE):
        return None

    weight = re.match(
        r"^([A-Za-zА-Яа-я][\w\s®™\-]{0,30}?)\s+(\d[\d.,\s]*(?:кг|kg|mm|мм|kW|W|л|l|%)[\w\s.]*)$",
        text,
        re.IGNORECASE,
    )
    if weight:
        return weight.group(1).strip(), weight.group(2).strip()

    branded_dash = re.match(
        r"^([A-Za-z0-9][A-Za-z0-9®™\-\s]{0,48}?)\s*[-–]\s*(.+)$",
        text,
    )
    if branded_dash:
        label = branded_dash.group(1).strip()
        value = branded_dash.group(2).strip()
        if value:
            return label, value

    any_dash = re.match(r"^([^\n]{2,55}?)\s*[-–]\s*(.+)$", text)
    if any_dash:
        label = any_dash.group(1).strip()
        value = any_dash.group(2).strip()
        if value and not re.match(r"^\d", label):
            return label, value

    colon = re.match(r"^([^:]{2,55}?):\s*(.+)$", text)
    if colon:
        label = colon.group(1).strip()
        value = colon.group(2).strip()
        if value and label.casefold() not in SKIP_SPEC_LABELS:
            return label, value

    period = text.find(". ")
    if period > 10:
        label = text[:period].strip()
        if len(label) <= 55:
            return label, text

    comma = text.find(",")
    if comma > 15:
        return text[:comma].strip(), text

    if "LED" in text.upper():
        return "LED осветление", text

    if any(keyword in text for keyword in FEATURE_KEYWORDS):
        words = text.split()
        label = " ".join(words[:6])
        return label, text

    words = text.split()
    if len(words) >= 2:
        if len(text) <= 50:
            return text, "Да"
        label = " ".join(words[:6])
        return label, text

    return None


def is_spec_candidate(label: str, value: str) -> bool:
    if not is_spec_label(label):
        return False
    has_digit = bool(re.search(r"\d", value))
    is_model = label.lower() in {"модел", "model"}
    has_degree = line_has_degree(value)
    return has_digit or is_model or has_degree


def split_description_specs(
    description: str,
) -> tuple[list[dict[str, str]], str]:
    """Pull label:value characteristic lines into specs and strip them from text."""
    found: list[dict[str, str]] = []
    seen: set[str] = set()
    kept_lines: list[str] = []

    for line in description.split("\n"):
        match = SPEC_LINE_RE.match(line.strip())
        if match:
            label = normalize_spec_label(match.group(1))
            value = match.group(2).strip().rstrip(".;")
            if is_spec_candidate(label, value):
                key = label.casefold()
                if key not in seen:
                    seen.add(key)
                    found.append({"label": label, "value": value})
                continue

        degree_spec = parse_degree_line(line)
        if degree_spec:
            label, value = degree_spec
            key = label.casefold()
            if key not in seen:
                seen.add(key)
                found.append({"label": label, "value": value})
            continue

        feature_spec = parse_feature_line(line)
        if feature_spec:
            label, value = feature_spec
            key = label.casefold()
            if key not in seen:
                seen.add(key)
                found.append({"label": label, "value": value})
            continue

        if is_feature_bullet_line(line):
            bullet_text = BULLET_LINE_RE.sub("", line.strip()).strip()
            if should_drop_feature_line(bullet_text):
                continue

        # Drop leftover section headers that only introduced the moved specs
        header = line.strip().rstrip(":").casefold()
        if header in SKIP_SPEC_LABELS or is_feature_section_header(line.strip()):
            continue

        kept_lines.append(line)

    cleaned = "\n".join(kept_lines)
    cleaned = re.sub(r"\n{3,}", "\n\n", cleaned).strip()
    # Drop trailing "download datasheet" parentheticals that often follow Модел
    cleaned = re.sub(
        r"\n?\(?(?:Изтеглете|За ПОДРОБНИ)[^)]*\)?\s*$",
        "",
        cleaned,
        flags=re.IGNORECASE,
    ).strip()
    return found, cleaned


I_FEATURE_LABEL_RE = re.compile(r"^i[A-Z]")
LONG_SPEC_VALUE_CHARS = 70


def is_i_feature(spec: dict[str, str]) -> bool:
    return bool(I_FEATURE_LABEL_RE.match(spec["label"]))


def is_long_spec(spec: dict[str, str]) -> bool:
    """Descriptive rows belong after short technical specs."""
    if is_i_feature(spec):
        return True
    return len(spec["value"].strip()) > LONG_SPEC_VALUE_CHARS


def sort_specs(specs: list[dict[str, str]]) -> list[dict[str, str]]:
    """Short technical specs first, then long text, i-features last."""
    short = [spec for spec in specs if not is_long_spec(spec)]
    long = [
        spec for spec in specs if is_long_spec(spec) and not is_i_feature(spec)
    ]
    i_features = [spec for spec in specs if is_i_feature(spec)]
    return short + long + i_features


def build_specs(
    row: tuple, description_specs: list[dict[str, str]] | None = None
) -> list[dict[str, str]]:
    specs: list[dict[str, str]] = []
    seen: set[str] = set()

    def add(label: str, value: str) -> None:
        key = label.casefold()
        if not value or key in seen:
            return
        seen.add(key)
        preserve_value = key == "модел" or preserves_leading_capitalization(value)
        preserve_label = preserves_leading_capitalization(label)
        clean_label = normalize_prose(label, preserve_leading=preserve_label)
        clean_value = normalize_prose(value, preserve_leading=preserve_value)
        if not clean_label or not clean_value:
            return
        # Placeholder / empty Excel cells
        if clean_value in {"-", "—", "–"}:
            return
        specs.append({"label": clean_label, "value": clean_value})

    power = clean_text(row[2]) if len(row) > 2 else ""
    origin = clean_text(row[3]) if len(row) > 3 else ""
    dimensions = clean_text(row[4]) if len(row) > 4 else ""
    if power:
        add("Захранване", power.replace("\n", " / "))
    if origin:
        add("Произход", origin)
    if dimensions:
        add("Размери (мм)", dimensions)

    for item in description_specs or []:
        add(item["label"], item["value"])

    return sort_specs(specs)


def extract_specs_from_description(description: str) -> list[dict[str, str]]:
    specs, _ = split_description_specs(description)
    return specs


def load_media_sizes(z: zipfile.ZipFile) -> dict[str, int]:
    sizes: dict[str, int] = {}
    for info in z.infolist():
        if info.filename.startswith("xl/media/"):
            sizes[Path(info.filename).name] = info.file_size
    return sizes


def load_sheet_images(
    z: zipfile.ZipFile, sheet_index: int
) -> dict[int, list[str]]:
    drawing_path = f"xl/drawings/drawing{sheet_index}.xml"
    rels_path = f"xl/drawings/_rels/drawing{sheet_index}.xml.rels"
    if drawing_path not in z.namelist():
        return {}

    rels_xml = z.read(rels_path).decode("utf-8")
    rel_map = {
        match.group(1): match.group(2)
        for match in re.finditer(
            r'Id="(rId\d+)"[^>]+Target="\.\./media/([^"]+)"', rels_xml
        )
    }

    root = ET.fromstring(z.read(drawing_path))
    by_row: dict[int, list[str]] = {}
    anchors = root.findall(".//xdr:twoCellAnchor", NS) + root.findall(
        ".//xdr:oneCellAnchor", NS
    )
    for anchor in anchors:
        from_el = anchor.find("xdr:from", NS)
        if from_el is None:
            continue
        row_el = from_el.find("xdr:row", NS)
        if row_el is None or row_el.text is None:
            continue
        row = int(row_el.text)
        blip = anchor.find(".//a:blip", NS)
        if blip is None:
            continue
        rid = blip.get(f"{{{NS['r']}}}embed")
        media = rel_map.get(rid or "")
        if media:
            by_row.setdefault(row, []).append(media)
    return by_row


def pick_product_image(
    excel_row: int, by_row: dict[int, list[str]], sizes: dict[str, int]
) -> str | None:
    candidates: list[str] = []
    for row_index in (excel_row - 2, excel_row - 1):
        candidates.extend(by_row.get(row_index, []))

    usable = [
        name
        for name in candidates
        if Path(name).suffix.lower() != ".emf" and sizes.get(name, 0) > 3000
    ]
    if usable:
        return max(usable, key=lambda name: sizes.get(name, 0))

    fallback = [
        name for name in candidates if Path(name).suffix.lower() != ".emf"
    ]
    if fallback:
        return max(fallback, key=lambda name: sizes.get(name, 0))
    return None


def extract_image(
    z: zipfile.ZipFile,
    media_name: str,
    slug: str,
) -> str | None:
    source = f"xl/media/{media_name}"
    if source not in z.namelist():
        return None

    ext = Path(media_name).suffix.lower()
    if ext == ".emf":
        return None

    dest_dir = IMAGES_ROOT / slug
    dest_dir.mkdir(parents=True, exist_ok=True)
    dest_file = dest_dir / f"1{ext}"
    dest_file.write_bytes(z.read(source))
    return f"/images/products/{slug}/1{ext}"


def ts_string(value: str) -> str:
    return json.dumps(value, ensure_ascii=False)


def render_ts(products: list[dict]) -> str:
    lines = [
        "/**",
        " * Imported from Catalogue-Webpage_data-Equipment-03.08.2026.xlsx",
        " * Regenerate: python scripts/import-catalogue.py",
        " */",
        'import type { Product } from "./schema";',
        "",
        "export const kitchenwareProducts: Product[] = [",
    ]

    for product in products:
        lines.append("  {")
        lines.append(f"    slug: {ts_string(product['slug'])},")
        lines.append(f"    subcategory: {ts_string(product['subcategory'])},")
        lines.append(f"    name: {ts_string(product['name'])},")
        lines.append(f"    tagline: {ts_string(product['tagline'])},")
        lines.append(f"    description: {ts_string(product['description'])},")
        if product.get("priceBgn"):
            lines.append(f"    priceBgn: {product['priceBgn']},")
        lines.append("    images: [")
        for image in product.get("images", []):
            lines.append(
                f"      {{ src: {ts_string(image['src'])}, alt: {ts_string(image['alt'])} }},"
            )
        lines.append("    ],")
        lines.append("    specs: [")
        for spec in product["specs"]:
            lines.append(
                f"      {{ label: {ts_string(spec['label'])}, value: {ts_string(spec['value'])} }},"
            )
        lines.append("    ],")
        lines.append(f"    status: {ts_string(product['status'])},")
        lines.append("  },")

    lines.append("];")
    lines.append("")
    return "\n".join(lines)


def parse_workbook(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    products: list[dict] = []
    seen_slugs: dict[str, int] = {}

    with zipfile.ZipFile(path) as z:
        media_sizes = load_media_sizes(z)

        for sheet_index, sheet_name in enumerate(wb.sheetnames, start=1):
            if sheet_index > len(SHEET_CONFIG):
                break
            config = SHEET_CONFIG[sheet_index - 1]
            ws = wb[sheet_name]
            rows = [tuple(row) for row in ws.iter_rows(values_only=True)]
            header_idx = find_header_row(rows)
            if header_idx is None:
                continue

            by_row = load_sheet_images(z, sheet_index)

            for row_index, row in enumerate(rows, start=1):
                if row_index <= header_idx + 1:
                    continue
                if not is_product_row(row):
                    continue

                raw = clean_text(row[1]) if len(row) > 1 else ""
                if not raw:
                    continue

                name = normalize_prose(product_name(raw))
                model = extract_model(raw)
                base_slug = slugify(model or name)
                count = seen_slugs.get(base_slug, 0)
                seen_slugs[base_slug] = count + 1
                slug = base_slug if count == 0 else f"{base_slug}-{count + 1}"

                price = None
                for col in (11, 12, 5, 7):
                    if len(row) > col:
                        price = parse_price(row[col])
                        if price:
                            break

                images: list[dict[str, str]] = []
                media_name = pick_product_image(row_index, by_row, media_sizes)
                if media_name:
                    src = extract_image(z, media_name, slug)
                    if src:
                        images.append({"src": src, "alt": name})

                desc_specs, description = split_description_specs(raw)

                products.append(
                    {
                        "slug": slug,
                        "subcategory": config["subcategory"],
                        "name": name,
                        "tagline": (
                            f"Модел {model} · {config['label']}"
                            if model
                            else config["label"]
                        ),
                        "description": normalize_prose(description),
                        "priceBgn": price,
                        "images": images,
                        "specs": build_specs(row, desc_specs),
                        "status": "published",
                        "sheet": sheet_name,
                    }
                )

    wb.close()
    return products


def main() -> None:
    source = DEFAULT_SOURCE if DEFAULT_SOURCE.exists() else FALLBACK_SOURCE
    if not source.exists():
        raise SystemExit(f"Catalogue file not found: {source}")

    products = parse_workbook(source)
    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUT_JSON.write_text(
        json.dumps(products, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    OUT_TS.write_text(render_ts(products), encoding="utf-8")

    with_images = sum(1 for product in products if product.get("images"))
    by_sub: dict[str, int] = {}
    for product in products:
        by_sub[product["subcategory"]] = by_sub.get(product["subcategory"], 0) + 1

    print(f"Imported {len(products)} products from {source.name}")
    print(f"Products with images: {with_images}/{len(products)}")
    for sub, count in by_sub.items():
        print(f"  {sub}: {count}")
    print(f"Wrote {OUT_TS.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
